"""
Exporta cada hoja de datos de KoraPay.xlsx a JSON normalizado en este directorio.
Transcripcion 1:1: cada fila con datos se convierte en un objeto JSON.
El seed.ts consume estos JSON. Correr una vez cuando cambie el Excel:

    python prisma/data/build.py

Requiere: openpyxl. Fuente por defecto:
  ../../../Recursos/korapay-personal/KoraPay.xlsx  (relativo a este archivo)
o pasar la ruta como primer argumento.
"""

import json
import os
import sys
import datetime
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
# data -> prisma -> korapay -> Codigo -> Mi-Bolsillito / Recursos / korapay-personal
DEFAULT_XLSX = os.path.normpath(
    os.path.join(HERE, "..", "..", "..", "..", "Recursos", "korapay-personal", "KoraPay.xlsx")
)
SRC = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

wb = openpyxl.load_workbook(SRC, data_only=True)


def iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    try:
        return round(float(str(v).replace(",", "").strip()), 6)
    except (ValueError, TypeError):
        return None


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def rows_from(sheet, header_row):
    ws = wb[sheet]
    out = []
    for r in ws.iter_rows(min_row=header_row + 1):
        cells = [c.value for c in r]
        if not any(c not in (None, "") for c in cells):
            continue
        out.append(cells)
    return out


def g(row, idx):
    return row[idx] if idx < len(row) else None


def write(name, records):
    path = os.path.join(HERE, f"{name}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=1)
    print(f"{name}.json: {len(records)} filas")
    return len(records)


# ---------------- Catalogos (hoja Menus) ----------------
def build_menus():
    ws = wb["Menús"]
    grid = [[c.value for c in r] for r in ws.iter_rows()]
    cols = {
        "empresas": 0,
        "medios_pago": 1,
        "monedas": 2,
        "tipo_pagos": 3,
        "tipos_movimiento": 4,
        "categorias_ingreso": 5,
        "categorias_gasto": 6,
        "categorias_fijos": 7,
        "meses": 8,
        "anios": 9,
        "personas_mimotech": 10,
        "suscripciones": 11,
    }
    out = {}
    for key, ci in cols.items():
        vals = []
        for r in range(1, len(grid)):
            v = grid[r][ci] if ci < len(grid[r]) else None
            t = text(v)
            if t and t not in vals:
                vals.append(t)
        out[key] = vals
    write("catalogs", out)


# ---------------- Personal ----------------
def build_ingresos_trabajos():
    recs = []
    for row in rows_from("IngresosM_Trabajos", 5):
        if not iso(g(row, 1)):
            continue
        recs.append(
            {
                "fecha": iso(g(row, 1)),
                "anio": num(g(row, 2)),
                "nMes": num(g(row, 3)),
                "mes": text(g(row, 4)),
                "trimestre": text(g(row, 5)),
                "tipo": text(g(row, 6)),
                "concepto": text(g(row, 7)),
                "empresa": text(g(row, 8)),
                "pago": text(g(row, 9)),
                "moneda": text(g(row, 10)) or "PEN",
                "totalSoles": num(g(row, 11)),
                "totalDolar": num(g(row, 12)),
                "totalNeto": num(g(row, 13)),
                "pagoEn": text(g(row, 14)),
                "numeroCuenta": text(g(row, 15)),
                "estado": text(g(row, 16)),
            }
        )
    write("ingresos_trabajos", recs)


def build_ingresos_empresas():
    recs = []
    for row in rows_from("IngresosM_Empresas", 5):
        if not iso(g(row, 1)):
            continue
        recs.append(
            {
                "fecha": iso(g(row, 1)),
                "anio": num(g(row, 2)),
                "nMes": num(g(row, 3)),
                "mes": text(g(row, 4)),
                "empresaOficial": text(g(row, 5)),
                "empresas": text(g(row, 6)),
                "empresasXMes": text(g(row, 7)),
                "fechaInicio": iso(g(row, 8)),
                "fechaFin": iso(g(row, 9)),
            }
        )
    write("ingresos_empresas", recs)


def build_egresos_personal():
    recs = []
    for row in rows_from("EgresosM_Personal", 5):
        if not iso(g(row, 1)):
            continue
        recs.append(
            {
                "fecha": iso(g(row, 1)),
                "anio": num(g(row, 2)),
                "nMes": num(g(row, 3)),
                "mes": text(g(row, 4)),
                "tipo": text(g(row, 5)),
                "fijoNoFijo": text(g(row, 6)),
                "concepto": text(g(row, 7)),
                "descripcion": text(g(row, 8)),
                "monto": num(g(row, 9)),
                "banco": text(g(row, 10)),
                "masDetalle": text(g(row, 11)),
                "estado": text(g(row, 12)),
            }
        )
    write("egresos_personal", recs)


def build_ahorros():
    recs = []
    for row in rows_from("InicioAhorroM", 5):
        if not iso(g(row, 1)):
            continue
        recs.append(
            {
                "fecha": iso(g(row, 1)),
                "anio": num(g(row, 2)),
                "nMes": num(g(row, 3)),
                "mes": text(g(row, 4)),
                "descripcion": text(g(row, 5)),
                "banco": text(g(row, 6)),
                "moneda": text(g(row, 7)) or "PEN",
                "monto": num(g(row, 8)),
                "importeTotal": num(g(row, 9)),
            }
        )
    write("ahorros", recs)


def build_renta():
    # IngresosM_Reporte: bloque Renta Anual (col Q=16 anio, R=17 monto, S=18 estado, T=19 detalles)
    ws = wb["IngresosM_Reporte"]
    grid = [[c.value for c in r] for r in ws.iter_rows()]
    recs = []
    for r in range(len(grid)):
        anio = grid[r][16] if len(grid[r]) > 16 else None
        monto = grid[r][17] if len(grid[r]) > 17 else None
        if isinstance(anio, (int, float)) and 2000 < anio < 2100 and monto is not None:
            recs.append(
                {
                    "anio": int(anio),
                    "monto": num(monto),
                    "estado": text(grid[r][18]) if len(grid[r]) > 18 else None,
                    "detalles": text(grid[r][19]) if len(grid[r]) > 19 else None,
                }
            )
    write("renta_anual", recs)


# ---------------- MIMOTECH ----------------
def build_mimotech_costos():
    recs = []
    for row in rows_from("Mimotech_Costos", 5):
        if not iso(g(row, 1)):
            continue
        recs.append(
            {
                "fecha": iso(g(row, 1)),
                "anio": num(g(row, 2)),
                "nMes": num(g(row, 3)),
                "mes": text(g(row, 4)),
                "aplicacion": text(g(row, 5)),
                "proyecto": text(g(row, 6)),
                "descripcion": text(g(row, 7)),
                "numeroTarjetaCuenta": text(g(row, 8)),
                "banco": text(g(row, 9)),
                "moneda": text(g(row, 10)) or "USD",
                "monto": num(g(row, 11)),
                "importeTotal": num(g(row, 12)),
                "estado": text(g(row, 13)),
            }
        )
    write("mimotech_costos", recs)


def build_mimotech_pagos():
    recs = []
    for row in rows_from("Mimotech_Pagos", 5):
        if not iso(g(row, 3)):
            continue
        recs.append(
            {
                "persona": text(g(row, 1)),
                "salario": text(g(row, 2)),
                "fecha": iso(g(row, 3)),
                "anio": num(g(row, 4)),
                "nMes": num(g(row, 5)),
                "mes": text(g(row, 6)),
                "estado": text(g(row, 7)),
                "notas": text(g(row, 8)),
                "monto": num(g(row, 9)),
            }
        )
    write("mimotech_pagos", recs)


# ---------------- Mimotalents ----------------
def build_talents_general():
    recs = []
    for row in rows_from("Mimotalents_General", 6):
        if not text(g(row, 1)):
            continue
        recs.append(
            {
                "nombre": text(g(row, 1)),
                "inicioConmigo": iso(g(row, 2)),
                "tiempoConmigo": text(g(row, 3)),
                "tiempoInicioPrimerTrabajo": text(g(row, 4)),
                "finConmigo": iso(g(row, 5)),
                "inicioPrimerTrabajo": iso(g(row, 6)),
                "tiempoFinContrato": text(g(row, 7)),
                "diapositiva": text(g(row, 8)),
                "lugarEstudio": text(g(row, 9)),
                "inicioEstudios": iso(g(row, 10)),
                "finEstudios": iso(g(row, 11)),
                "estado": text(g(row, 12)),
            }
        )
    write("talents_general", recs)


def build_talents_ingresos():
    recs = []
    for row in rows_from("Mimotalents_Ingresos", 5):
        if not text(g(row, 1)) or not iso(g(row, 2)):
            continue
        recs.append(
            {
                "nombre": text(g(row, 1)),
                "fecha": iso(g(row, 2)),
                "anio": num(g(row, 3)),
                "nMes": num(g(row, 4)),
                "mes": text(g(row, 5)),
                "tipoPago": text(g(row, 6)),
                "empresa": text(g(row, 7)),
                "cliente": text(g(row, 8)),
                "pagos": text(g(row, 9)),
                "cargo": text(g(row, 10)),
                "sueldo": num(g(row, 11)),
                "conDescuento": num(g(row, 12)),
                "recibi": num(g(row, 13)),
                "seQuedoCon": num(g(row, 14)),
                "estado": text(g(row, 15)),
                "tiempoContrato": text(g(row, 16)),
                "duro": text(g(row, 17)),
                "inicio": iso(g(row, 18)),
                "fin": iso(g(row, 19)),
            }
        )
    write("talents_ingresos", recs)


def build_talents_egresos():
    recs = []
    for row in rows_from("Mimotalents_Egresos", 5):
        if not text(g(row, 1)) or not iso(g(row, 2)):
            continue
        recs.append(
            {
                "nombre": text(g(row, 1)),
                "fecha": iso(g(row, 2)),
                "anio": num(g(row, 3)),
                "nMes": num(g(row, 4)),
                "mes": text(g(row, 5)),
                "tipoPago": text(g(row, 6)),
                "cantidadE": num(g(row, 7)),
                "cantidadD": num(g(row, 8)),
                "faltaPagar": num(g(row, 9)),
                "descripcion": text(g(row, 10)),
                "estado": text(g(row, 11)),
            }
        )
    write("talents_egresos", recs)


def main():
    build_menus()
    build_ingresos_trabajos()
    build_ingresos_empresas()
    build_egresos_personal()
    build_ahorros()
    build_renta()
    build_mimotech_costos()
    build_mimotech_pagos()
    build_talents_general()
    build_talents_ingresos()
    build_talents_egresos()
    print("OK. Fuente:", SRC)


if __name__ == "__main__":
    main()
